#!/usr/bin/env python3
"""Normalize the legacy Hà Hoà workbook into relational import records."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel


NAMESPACE = uuid.UUID("d60f98f4-a81f-4a83-85a7-1125a2ea3f45")
MAX_IMPORT_ROWS = 2500
MAX_IMPORT_COLS = 20


def normalized_text(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", text.replace("đ", "d"))


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def number(value: Any) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        result = float(value)
        return result if math.isfinite(result) else None
    text = str(value).strip()
    if text.startswith("=") or text.startswith("#"):
        return None
    check_text = normalized_text(text).replace("vnd", "").replace("dong", "")
    if re.search(r"[a-z]", check_text):
        return None
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text:
        return None
    if text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "").replace(".", "")
    try:
        result = float(text)
        return result if math.isfinite(result) else None
    except ValueError:
        return None


def payment_amount(value: Any, debt_amount: float) -> float | None:
    parsed = number(value)
    if parsed is not None:
        return parsed
    if not isinstance(value, str):
        return None
    text = normalized_text(value)
    remaining_match = re.search(r"con no(?: lai)?\s*([0-9][0-9.,]*)", text)
    if remaining_match:
        remaining = number(remaining_match.group(1))
        if remaining is not None and 0 <= remaining <= debt_amount:
            return debt_amount - remaining
    paid_match = re.search(r"(?:da tra|tra|tru)\s*([0-9][0-9.,]*)", text)
    if paid_match:
        paid = number(paid_match.group(1))
        if paid is not None and paid < 1000 and debt_amount >= 10_000:
            paid *= 1000
        if paid is not None and 0 < paid <= debt_amount:
            return paid
    return None


def excel_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 30_000 <= value <= 80_000:
        try:
            parsed = from_excel(value)
            return parsed.date() if isinstance(parsed, datetime) else parsed
        except (OverflowError, ValueError):
            return None
    text = str(value).strip()
    text = text.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed if 1990 <= parsed.year <= 2100 else None
        except ValueError:
            continue
    return None


def iso(value: date | None) -> str | None:
    return value.isoformat() if value else None


def classify_header(value: Any) -> str | None:
    text = normalized_text(value)
    if not text:
        return None
    if "ma kh" in text:
        return "customer_code"
    if "ten khach hang" in text:
        return "customer_name"
    if "so tien no" in text:
        return "amount"
    if text in {"nvkd no", "nhan vien kd no", "nv phu trach", "nguoi chiu trach nhiem"}:
        return "sales_person"
    if "nv giao" in text or "nhan vien giao" in text:
        return "delivery_person"
    if "ngay don hang" in text or "ngay giao hang" in text or text == "ngay no":
        return "order_date"
    if "so tien da tra" in text or "da thanh toan" in text:
        return "paid_amount"
    if text in {"ngay thanh toan", "ngay tt", "ngay tra"}:
        return "payment_date"
    if "han thanh toan" in text or text == "han no" or text == "han tra no":
        return "due_term"
    if text == "ghi chu":
        return "notes"
    return None


def infer_sheet_date(rows: list[tuple[Any, ...]]) -> date:
    heading = " ".join(str(cell or "") for row in rows[:5] for cell in row[:10])
    match = re.search(r"th[aá]ng\s*([01]?\d)\s*(?:/|n[aă]m)?\s*(20\d{2})", heading, re.IGNORECASE)
    if match:
        month, year = int(match.group(1)), int(match.group(2))
        if 1 <= month <= 12:
            return date(year, month, 1)
    return date(2026, 8, 1)


def due_days(value: Any, order_date: date) -> int:
    parsed_date = excel_date(value)
    if parsed_date:
        delta = (parsed_date - order_date).days
        return delta if 0 <= delta <= 3650 else 30
    parsed_number = number(value)
    if parsed_number is not None and 0 <= parsed_number <= 3650:
        return int(parsed_number)
    return 30


def parse(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    customers: dict[str, dict[str, Any]] = {}
    debts: list[dict[str, Any]] = []
    payments: list[dict[str, Any]] = []
    skipped_sheets: list[str] = []
    inferred_order_dates = 0
    inferred_payment_dates = 0

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, MAX_IMPORT_ROWS), min_col=1, max_col=min(sheet.max_column, MAX_IMPORT_COLS), values_only=True))
        header_row = None
        columns: dict[str, int] = {}
        for row_index, row in enumerate(rows[:12], start=1):
            candidate = {field: index for index, value in enumerate(row) if (field := classify_header(value))}
            if "customer_name" in candidate and "amount" in candidate:
                header_row, columns = row_index, candidate
                break
        if not header_row:
            skipped_sheets.append(sheet.title)
            continue

        fallback_date = infer_sheet_date(rows)
        for source_row, row in enumerate(rows[header_row:], start=header_row + 1):
            customer_name = clean_text(row[columns["customer_name"]])
            amount = number(row[columns["amount"]])
            if not customer_name or amount is None or amount <= 0:
                continue

            customer_key = normalized_text(customer_name)
            customer_id = str(uuid.uuid5(NAMESPACE, f"customer:{customer_key}"))
            customer_code = clean_text(row[columns["customer_code"]]) if "customer_code" in columns else None
            if customer_key not in customers:
                customers[customer_key] = {"id": customer_id, "name": customer_name, "code": customer_code}
            elif not customers[customer_key].get("code") and customer_code:
                customers[customer_key]["code"] = customer_code

            order_value = row[columns["order_date"]] if "order_date" in columns else None
            order_date = excel_date(order_value)
            if not order_date:
                order_date = fallback_date
                inferred_order_dates += 1

            debt_id = str(uuid.uuid5(NAMESPACE, f"debt:{sheet.title}:{source_row}"))
            notes = clean_text(row[columns["notes"]]) if "notes" in columns else None
            paid_value = row[columns["paid_amount"]] if "paid_amount" in columns else None
            if isinstance(paid_value, str) and paid_value.strip():
                payment_marker = f"Nội dung cột thanh toán Excel: {paid_value.strip()}"
                notes = f"{notes} · {payment_marker}" if notes else payment_marker
            debt = {
                "id": debt_id,
                "customer_id": customer_id,
                "amount": amount,
                "order_date": iso(order_date),
                "due_days": due_days(row[columns["due_term"]], order_date) if "due_term" in columns else 30,
                "sales_person": clean_text(row[columns["sales_person"]]) if "sales_person" in columns else None,
                "delivery_person": clean_text(row[columns["delivery_person"]]) if "delivery_person" in columns else None,
                "notes": notes,
                "source_sheet": sheet.title,
                "source_row": source_row,
            }
            debts.append(debt)

            paid_amount = payment_amount(paid_value, amount) if "paid_amount" in columns else None
            if paid_amount and paid_amount > 0:
                payment_value = row[columns["payment_date"]] if "payment_date" in columns else None
                payment_date = excel_date(payment_value)
                payment_notes = notes
                if not payment_date:
                    payment_date = order_date
                    inferred_payment_dates += 1
                    marker = "Ngày trả được tạm lấy theo ngày nợ khi import Excel."
                    payment_notes = f"{notes} · {marker}" if notes else marker
                payments.append({
                    "id": str(uuid.uuid5(NAMESPACE, f"payment:{sheet.title}:{source_row}")),
                    "debt_id": debt_id,
                    "amount": paid_amount,
                    "paid_at": iso(payment_date),
                    "sales_person": debt["sales_person"],
                    "delivery_person": debt["delivery_person"],
                    "notes": payment_notes,
                    "source_sheet": sheet.title,
                    "source_row": source_row,
                })

    result = {
        "customers": list(customers.values()),
        "debts": debts,
        "payments": payments,
        "stats": {
            "source": path.name,
            "sheets_total": len(workbook.sheetnames),
            "sheets_imported": len(workbook.sheetnames) - len(skipped_sheets),
            "sheets_skipped": skipped_sheets,
            "customers": len(customers),
            "debts": len(debts),
            "payments": len(payments),
            "inferred_order_dates": inferred_order_dates,
            "inferred_payment_dates": inferred_payment_dates,
        },
    }
    workbook.close()
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--stats", action="store_true")
    args = parser.parse_args()
    if not args.workbook.exists():
        raise SystemExit(f"Không tìm thấy workbook: {args.workbook}")
    result = parse(args.workbook)
    json.dump(result["stats"] if args.stats else result, sys.stdout, ensure_ascii=False, separators=(",", ":"))
    sys.stdout.write("\n")


if __name__ == "__main__":
    main()
